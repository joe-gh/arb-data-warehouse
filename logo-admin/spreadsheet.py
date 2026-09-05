"""Defensive spreadsheet parsing and owner-scoped two-confirmation jobs."""

import csv
import asyncio
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal
import hashlib
import io
import json
import os
from pathlib import Path, PurePosixPath
import re
import shutil
import stat
import uuid
import zipfile
from typing import Any, Mapping, Optional
import xml.etree.ElementTree as ET

import openpyxl
from psycopg2.extras import Json

from commands import (
    COMMAND_MODELS,
    MutationCommand,
    SaveAssignmentCommand,
    SetStorePricingTierCommand,
)
from db import database
from domain import Conflict, InvalidCommand, NotFound
from spreadsheet_mapping import (
    MappingCapacityExceeded,
    SHEET_COMMAND_NAMES,
    TARGET_FIELDS,
    MappingProposal,
    propose_mapping,
    SpreadsheetNameResolver,
    NameResolutionError,
    validate_mapping_headers,
)
import staging
import quotas


ASSIGNMENT_COLUMNS = (
    "fdm4_store",
    "product_style",
    "garment_color_code",
    "option_row",
    "position",
    "design_id",
    "logo_code",
    "color_scheme_id",
    "location",
    "optional",
    "background",
    "cost_override",
    "sort_order",
    "image_url",
    "active",
)
PRICING_COLUMNS = ("fdm4_store", "tier_name", "note")
MAX_ACTIVE_SPREADSHEET_JOBS_PER_USER = 10
MAX_ACTIVE_SPREADSHEET_BYTES_MULTIPLIER = 5
MAX_ACTIVE_SPREADSHEET_JOBS_GLOBAL = 100
MAX_ACTIVE_SPREADSHEET_BYTES_GLOBAL_MULTIPLIER = 50


@dataclass(frozen=True)
class SpreadsheetLimits:
    max_bytes: int = 5 * 1024 * 1024
    max_rows: int = 500
    max_columns: int = 40
    max_cell_chars: int = 2000
    max_xlsx_entries: int = 200
    max_xlsx_uncompressed_bytes: int = 50 * 1024 * 1024
    max_compression_ratio: int = 100

    @classmethod
    def from_settings(cls, settings) -> "SpreadsheetLimits":
        return cls(
            max_bytes=settings.agent_max_spreadsheet_bytes,
            max_rows=settings.agent_max_spreadsheet_rows,
            max_columns=settings.agent_max_spreadsheet_columns,
            max_cell_chars=settings.agent_max_cell_chars,
            max_xlsx_entries=settings.agent_max_xlsx_entries,
            max_xlsx_uncompressed_bytes=settings.agent_max_xlsx_uncompressed_bytes,
        )


@dataclass(frozen=True)
class ParsedSpreadsheet:
    format_name: str
    headers: tuple[str, ...]
    rows: tuple[dict[str, Any], ...]


async def _joinable_to_thread(function, /, *args, **kwargs):
    task = asyncio.create_task(asyncio.to_thread(function, *args, **kwargs))
    try:
        return await asyncio.shield(task)
    except asyncio.CancelledError:
        try:
            await asyncio.shield(task)
        except Exception:
            pass
        raise


async def _reserve_job_cancellation_safe(
    *,
    session_id,
    user_login: str,
    data: bytes,
    filename: str,
    media_type: str,
    settings,
) -> dict:
    task = asyncio.create_task(asyncio.to_thread(
        _reserve_job,
        session_id=session_id,
        user_login=user_login,
        data=data,
        filename=filename,
        media_type=media_type,
        settings=settings,
    ))
    try:
        return await asyncio.shield(task)
    except asyncio.CancelledError:
        try:
            job = await asyncio.shield(task)
        except Exception:
            pass
        else:
            await asyncio.shield(asyncio.to_thread(
                _discard_reserved_job,
                job,
                user_login,
                settings,
            ))
        raise


def _normalize_header(value: Any) -> str:
    header = "_".join(str(value or "").strip().lower().split())
    if not header or "\x00" in header or len(header) > 100:
        raise InvalidCommand("Spreadsheet contains an invalid header")
    return header


def _cell(value: Any, limits: SpreadsheetLimits) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool, Decimal)):
        rendered = value if not isinstance(value, str) else value.strip()
    elif hasattr(value, "isoformat"):
        rendered = value.isoformat()
    else:
        rendered = str(value)
    if len(str(rendered)) > limits.max_cell_chars or "\x00" in str(rendered):
        raise InvalidCommand("Spreadsheet cell exceeds the supported limit")
    return rendered


def _rows_from_matrix(
    matrix: list[list[Any]],
    limits: SpreadsheetLimits,
) -> tuple[tuple[str, ...], tuple[dict[str, Any], ...]]:
    if not matrix:
        raise InvalidCommand("Spreadsheet is empty")
    if len(matrix[0]) > limits.max_columns:
        raise InvalidCommand("Spreadsheet has too many columns")
    headers = tuple(_normalize_header(value) for value in matrix[0])
    if len(set(headers)) != len(headers):
        raise InvalidCommand("Spreadsheet headers must be unique")
    output = []
    for line, values in enumerate(matrix[1:], start=2):
        if line - 1 > limits.max_rows:
            raise InvalidCommand("Spreadsheet has too many rows")
        if len(values) > len(headers):
            raise InvalidCommand(f"Row {line} has more cells than headers")
        normalized = [
            _cell(value, limits)
            for value in values + ([""] * (len(headers) - len(values)))
        ]
        if not any(str(value).strip() for value in normalized):
            continue
        output.append(dict(zip(headers, normalized)))
    if not output:
        raise InvalidCommand("Spreadsheet has no data rows")
    return headers, tuple(output)


def _parse_csv(data: bytes, limits: SpreadsheetLimits) -> ParsedSpreadsheet:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise InvalidCommand("CSV must be UTF-8 encoded") from None
    if "\x00" in text:
        raise InvalidCommand("CSV contains a NUL byte")
    matrix: list[list[str]] = []
    try:
        reader = csv.reader(io.StringIO(text, newline=""), strict=True)
        for row_number, row in enumerate(reader, start=1):
            if row_number > limits.max_rows + 1:
                raise InvalidCommand("Spreadsheet has too many rows")
            if len(row) > limits.max_columns:
                raise InvalidCommand("Spreadsheet has too many columns")
            matrix.append(list(row))
    except csv.Error as exc:
        raise InvalidCommand(f"Malformed CSV: {exc}") from None
    headers, rows = _rows_from_matrix(matrix, limits)
    return ParsedSpreadsheet("csv", headers, rows)


def _inspect_xlsx(data: bytes, limits: SpreadsheetLimits) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        raise InvalidCommand("XLSX is not a valid ZIP container") from None
    with archive:
        entries = archive.infolist()
        if len(entries) > limits.max_xlsx_entries:
            raise InvalidCommand("XLSX has too many archive entries")
        total = 0
        seen_paths: set[str] = set()
        for entry in entries:
            path = PurePosixPath(entry.filename)
            normalized_path = entry.filename.lower()
            if (
                path.is_absolute()
                or ".." in path.parts
                or "\\" in entry.filename
                or "\x00" in entry.filename
                or normalized_path in seen_paths
            ):
                raise InvalidCommand("XLSX contains an unsafe archive path")
            seen_paths.add(normalized_path)
            if entry.flag_bits & 0x1:
                raise InvalidCommand("Encrypted XLSX entries are not supported")
            if (
                normalized_path == "xl/vbaproject.bin"
                or normalized_path == "xl/connections.xml"
                or normalized_path.startswith((
                    "xl/externallinks/",
                    "xl/activex/",
                    "xl/embeddings/",
                    "xl/querytables/",
                ))
            ):
                raise InvalidCommand(
                    "Macros, external links, and embedded content are not supported"
                )
            total += entry.file_size
            if total > limits.max_xlsx_uncompressed_bytes:
                raise InvalidCommand("XLSX expands beyond the supported limit")
            if entry.file_size and entry.compress_size == 0:
                raise InvalidCommand("XLSX has an invalid compression entry")
            if entry.compress_size and entry.file_size / entry.compress_size > limits.max_compression_ratio:
                raise InvalidCommand("XLSX compression ratio is unsafe")

            if not entry.is_dir() and normalized_path.endswith(".rels"):
                relationships = archive.read(entry)
                if re.search(
                    rb"targetmode\s*=\s*['\"]external['\"]",
                    relationships.lower(),
                ):
                    raise InvalidCommand("External XLSX relationships are not supported")

            if (
                not entry.is_dir()
                and normalized_path.startswith("xl/worksheets/")
                and normalized_path.endswith(".xml")
            ):
                worksheet_xml = archive.read(entry)
                lowered_xml = worksheet_xml.lower()
                if b"<!doctype" in lowered_xml or b"<!entity" in lowered_xml:
                    raise InvalidCommand("XLSX worksheet XML contains a prohibited DTD")
                row_count = 0
                try:
                    for _event, element in ET.iterparse(
                        io.BytesIO(worksheet_xml),
                        events=("end",),
                    ):
                        local_name = str(element.tag).rsplit("}", 1)[-1]
                        if local_name == "row":
                            row_count += 1
                            if row_count > limits.max_rows + 1:
                                raise InvalidCommand("Spreadsheet has too many rows")
                        elif local_name == "c":
                            reference = str(element.attrib.get("r", ""))
                            match = re.match(r"^([A-Za-z]+)", reference)
                            if match:
                                column = 0
                                for character in match.group(1).upper():
                                    column = column * 26 + ord(character) - 64
                                if column > limits.max_columns:
                                    raise InvalidCommand("Spreadsheet has too many columns")
                        elif local_name == "f":
                            raise InvalidCommand("Spreadsheet formulas are not supported")
                        element.clear()
                except ET.ParseError:
                    raise InvalidCommand("XLSX contains malformed worksheet XML") from None


def _parse_xlsx(data: bytes, limits: SpreadsheetLimits) -> ParsedSpreadsheet:
    _inspect_xlsx(data, limits)
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise InvalidCommand(f"Unable to read XLSX: {exc}") from None
    try:
        if not workbook.worksheets:
            raise InvalidCommand("XLSX has no worksheets")
        worksheet = workbook.worksheets[0]
        if worksheet.max_row > limits.max_rows + 1:
            raise InvalidCommand("Spreadsheet has too many rows")
        if worksheet.max_column > limits.max_columns:
            raise InvalidCommand("Spreadsheet has too many columns")
        matrix: list[list[Any]] = []
        # Iterate the sheet's actual width (already validated <= max_columns).
        # iter_rows(max_col=limits.max_columns) pads every row out to that width
        # with None cells, which then fails header normalization for any sheet
        # narrower than max_columns (i.e. essentially all real spreadsheets).
        column_span = worksheet.max_column or limits.max_columns
        for row_index, row in enumerate(
            worksheet.iter_rows(
                max_row=limits.max_rows + 1,
                max_col=column_span,
            ),
            start=1,
        ):
            if row_index > limits.max_rows + 1:
                raise InvalidCommand("Spreadsheet has too many rows")
            if len(row) > limits.max_columns:
                raise InvalidCommand("Spreadsheet has too many columns")
            values = []
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.lstrip().startswith("=")
                ):
                    raise InvalidCommand("Spreadsheet formulas are not supported")
                values.append(cell.value)
            matrix.append(values)
    finally:
        workbook.close()
    headers, rows = _rows_from_matrix(matrix, limits)
    return ParsedSpreadsheet("xlsx", headers, rows)


def parse_spreadsheet(
    data: bytes,
    filename: str,
    limits: SpreadsheetLimits,
) -> ParsedSpreadsheet:
    if not data:
        raise InvalidCommand("Spreadsheet is empty")
    if len(data) > limits.max_bytes:
        raise InvalidCommand("Spreadsheet exceeds the upload limit")
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        return _parse_csv(data, limits)
    if suffix == ".xlsx":
        return _parse_xlsx(data, limits)
    raise InvalidCommand("Only .csv and .xlsx files are supported")


def known_mapping(parsed: ParsedSpreadsheet) -> Optional[MappingProposal]:
    headers = set(parsed.headers)
    if "command" in headers:
        return MappingProposal(command="mixed", columns={field: field for field in TARGET_FIELDS["mixed"] if field in headers})
    if set(ASSIGNMENT_COLUMNS).issubset(headers):
        return MappingProposal(
            command="save_assignment",
            columns={field: field for field in ASSIGNMENT_COLUMNS},
        )
    if set(PRICING_COLUMNS).issubset(headers):
        return MappingProposal(
            command="set_store_pricing_tier",
            columns={field: field for field in PRICING_COLUMNS},
        )
    return None



def _sheet_command_counts(parsed, proposal):
    counts = {}
    for row in parsed.rows:
        name = proposal.command
        if name == "mixed":
            name = str(row.get(proposal.columns.get("command"),proposal.constants.get("command",""))).strip()
        label = name if name in SHEET_COMMAND_NAMES else "unsupported rows"
        counts[label] = counts.get(label,0)+1
    return counts

def mapping_hash(proposal: MappingProposal) -> str:
    raw = json.dumps(
        proposal.model_dump(mode="json"),
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _bool(value: Any, default: bool) -> bool:
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    raise ValueError("must be true or false")


def _integer(value: Any, field: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{field} must be an integer")
    try:
        parsed = Decimal(str(value).strip())
    except Exception:
        raise ValueError(f"{field} must be an integer") from None
    if not parsed.is_finite() or parsed != parsed.to_integral_value():
        raise ValueError(f"{field} must be an integer")
    return int(parsed)


def _restore_csv_text(value: Any, *, csv_format: bool) -> Any:
    if not csv_format:
        return value
    rendered = "" if value is None else str(value)
    if rendered.startswith("''"):
        return rendered[1:]
    if rendered.startswith("'"):
        candidate = rendered[1:]
        probe = candidate.lstrip(" \t\r\n")
        if candidate.startswith(("\t", "\r", "\n")) or probe.startswith(("=", "+", "-", "@")):
            return candidate
    return value


def _translate_rows_with_numbers(
    parsed: ParsedSpreadsheet,
    proposal: MappingProposal,
    *, resolver=None, resolutions=None,
) -> tuple[list[tuple[int, MutationCommand]], list[dict]]:
    validate_mapping_headers(proposal, parsed.headers)
    commands: list[tuple[int, MutationCommand]] = []
    rejected: list[dict] = []
    csv_format = parsed.format_name == "csv"
    for index, row in enumerate(parsed.rows, start=2):
        values = dict(proposal.constants)
        values.update({
            target: _restore_csv_text(
                row.get(source, ""),
                csv_format=csv_format,
            )
            for target, source in proposal.columns.items()
        })
        try:
            tool_name = str(values.pop("command", "")).strip() if proposal.command == "mixed" else proposal.command
            if tool_name not in SHEET_COMMAND_NAMES:
                raise ValueError("Unsupported spreadsheet command")
            if proposal.command == "mixed":
                values = {k: v for k, v in values.items() if k in TARGET_FIELDS[tool_name] and v not in ("", None)}
            row_resolutions = []
            if resolver is not None:
                # Reuse exact name resolution with the field names its lookups own.
                aliases = {"store": "fdm4_store", "style_code": "product_style"}
                adapted = {aliases.get(k,k): v for k,v in values.items() if k != "styles"}
                # A global stock-override removal needs a code, not a store-dependent name guess.
                if tool_name == "remove_stock_override":
                    adapted.pop("product_style",None)
                adapted, row_resolutions = resolver.resolve(adapted, index)
                for key in list(values):
                    target = aliases.get(key,key)
                    if target in adapted:
                        values[key] = adapted[target]
            if tool_name == "save_assignment":
                values.setdefault("option_row", 1)
                values.setdefault("location", "")
                values.setdefault("optional", False)
                values.setdefault("background", "")
                values.setdefault("cost_override", None)
                values.setdefault("sort_order", 0)
                values.setdefault("image_url", "")
                values.setdefault("active", True)
                values["optional"] = _bool(values.get("optional"), False)
                values["active"] = _bool(values.get("active"), True)
                for key in ("option_row", "position", "sort_order"):
                    values[key] = _integer(values[key], key)
                if values.get("cost_override") in {"", None}:
                    values["cost_override"] = None
                command: MutationCommand = SaveAssignmentCommand.model_validate(values)
            elif tool_name == "set_store_pricing_tier":
                values.setdefault("note", "")
                command = SetStorePricingTierCommand.model_validate(values)
            else:
                if tool_name == "deactivate_assignment":
                    values["option_row"] = _integer(values.get("option_row",1),"option_row")
                    values["position"] = _integer(values.get("position"),"position")
                if tool_name == "delete_price_rule":
                    values["rule_id"] = _integer(values.get("rule_id"),"rule_id")
                if "styles" in TARGET_FIELDS[tool_name]:
                    raw_styles = str(values.get("styles","")).strip()
                    if proposal.command == "mixed" and tool_name == "remove_sync_block":
                        # In a mixed sheet a blank cell is an omission; the whole store must be asked for.
                        if raw_styles == "*":
                            raw_styles = ""
                        elif not raw_styles:
                            raise ValueError("Name the styles to unfreeze, or write * for the whole store")
                    styles = [v.strip() for v in raw_styles.split(",") if v.strip()]
                    if len(styles)>50:
                        raise ValueError("At most 50 styles per row")
                    if resolver is not None:
                        resolved_styles = []
                        for style in styles:
                            resolved, report = resolver.resolve({"fdm4_store":values.get("store"),"product_style":style},index)
                            resolved_styles.append(resolved["product_style"])
                            row_resolutions.extend(report)
                        styles = resolved_styles
                    values["styles"] = styles
                command = COMMAND_MODELS[tool_name].model_validate(values)
            commands.append((index, command))
            if resolutions is not None:
                resolutions.extend(row_resolutions)
        except Exception as exc:
            error = {"row": index, "detail": str(exc)[:500]}
            if isinstance(exc, NameResolutionError):
                error.update({"field": exc.field, "candidates": exc.candidates})
            rejected.append(error)
    return commands, rejected


def translate_rows(
    parsed: ParsedSpreadsheet,
    proposal: MappingProposal,
) -> tuple[list[MutationCommand], list[dict]]:
    """Public pure translation contract retained for tests and callers."""

    numbered, rejected = _translate_rows_with_numbers(parsed, proposal)
    return [command for _row, command in numbered], rejected


def _job_path(settings, storage_key: uuid.UUID) -> Path:
    try:
        safe_key = uuid.UUID(str(storage_key))
    except (ValueError, TypeError, AttributeError):
        raise InvalidCommand("Invalid private upload key") from None
    base = Path(settings.agent_upload_dir)
    candidate = base / f"{safe_key}.upload"
    if candidate.parent != base:
        raise InvalidCommand("Invalid private upload path")
    return candidate


def _session_uuid(value) -> uuid.UUID:
    try:
        return uuid.UUID(str(value))
    except (ValueError, TypeError, AttributeError):
        raise NotFound("Chat session not found") from None


def _remove_private_file(path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    except OSError:
        # The retention sweep retries safe UUID-named files. A cleanup failure
        # must not roll back an already-staged database change-set.
        pass


def _format_hint(filename: str) -> str:
    suffix = Path(filename.replace("\\", "/")).suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    raise InvalidCommand("Spreadsheet must be a .csv or .xlsx file")


def _reserve_job(
    *,
    session_id,
    user_login: str,
    data: bytes,
    filename: str,
    media_type: str,
    settings,
) -> dict:
    """Reserve aggregate capacity and private storage before doing work."""

    digest = hashlib.sha256(data).hexdigest()
    job_id = uuid.uuid4()
    storage_key = uuid.uuid4()
    format_name = _format_hint(filename)
    path = _job_path(settings, storage_key)
    path.parent.mkdir(mode=0o700, parents=True, exist_ok=True)
    path.parent.chmod(0o700)
    if shutil.disk_usage(path.parent).free < (
        len(data) + settings.agent_max_spreadsheet_bytes
    ):
        raise InvalidCommand("Private spreadsheet storage is too full")
    persisted = None
    try:
        with database.cursor(write=True, actor=user_login) as cursor:
            cursor.execute(
                """
                SELECT 1 FROM logo.agent_chat_session
                 WHERE id = %s AND user_login = %s AND expires_at > now()
                """,
                (_session_uuid(session_id), user_login),
            )
            if cursor.fetchone() is None:
                raise NotFound("Chat session not found")
            for lock_name in sorted((
                "agent-spreadsheet-upload:global",
                f"agent-spreadsheet-upload:{user_login}",
            )):
                cursor.execute(
                    "SELECT pg_advisory_xact_lock(hashtextextended(%s, 0))",
                    (lock_name,),
                )
            cursor.execute(
                """
                SELECT count(*) AS jobs, coalesce(sum(byte_size), 0) AS bytes
                 FROM logo.agent_spreadsheet_job
                 WHERE user_login = %s AND expires_at > now()
                   AND status IN (
                       'mapping_processing',
                       'mapping_pending',
                       'mapping_confirmed'
                   )
                """,
                (user_login,),
            )
            active = cursor.fetchone()
            if (
                int(active["jobs"]) >= MAX_ACTIVE_SPREADSHEET_JOBS_PER_USER
                or int(active["bytes"]) + len(data)
                > settings.agent_max_spreadsheet_bytes
                * MAX_ACTIVE_SPREADSHEET_BYTES_MULTIPLIER
            ):
                raise InvalidCommand(
                    "Too many active spreadsheet uploads; finish or wait for cleanup"
                )
            cursor.execute(
                """
                SELECT count(*) AS jobs, coalesce(sum(byte_size), 0) AS bytes
                 FROM logo.agent_spreadsheet_job
                 WHERE expires_at > now()
                   AND status IN (
                       'mapping_processing',
                       'mapping_pending',
                       'mapping_confirmed'
                   )
                """
            )
            global_active = cursor.fetchone()
            if (
                int(global_active["jobs"])
                >= MAX_ACTIVE_SPREADSHEET_JOBS_GLOBAL
                or int(global_active["bytes"]) + len(data)
                > settings.agent_max_spreadsheet_bytes
                * MAX_ACTIVE_SPREADSHEET_BYTES_GLOBAL_MULTIPLIER
            ):
                raise InvalidCommand(
                    "Private spreadsheet storage capacity is reached"
                )
            cursor.execute(
                """
                INSERT INTO logo.agent_spreadsheet_job (
                    id, session_id, user_login, storage_key, original_name,
                    media_type, byte_size, sha256, format_name, status,
                    mapping_revision, mapping_hash, mapping, expires_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    'mapping_processing', 1, %s, '{}'::jsonb, %s
                ) RETURNING *
                """,
                (
                    job_id,
                    _session_uuid(session_id),
                    user_login,
                    storage_key,
                    Path(filename.replace("\\", "/")).name[:255],
                    str(media_type or "application/octet-stream")[:255],
                    len(data),
                    digest,
                    format_name,
                    "0" * 64,
                    datetime.now(timezone.utc) + timedelta(hours=1),
                ),
            )
            persisted = dict(cursor.fetchone())

        # The committed metadata is the reservation: concurrent requests count
        # it before any upload bytes are written. A crash leaves a bounded job
        # whose missing file is safely rejected/cleaned.
        descriptor = os.open(
            path,
            os.O_WRONLY | os.O_CREAT | os.O_EXCL
            | getattr(os, "O_NOFOLLOW", 0),
            0o600,
        )
        try:
            with os.fdopen(descriptor, "wb") as private_file:
                descriptor = None
                private_file.write(data)
        finally:
            if descriptor is not None:
                os.close(descriptor)
        return persisted
    except Exception:
        _remove_private_file(path)
        if persisted is not None:
            try:
                with database.cursor(write=True, actor=user_login) as cursor:
                    cursor.execute(
                        """
                        DELETE FROM logo.agent_spreadsheet_job
                         WHERE id = %s AND user_login = %s
                           AND status = 'mapping_processing'
                        """,
                        (job_id, user_login),
                    )
            except Exception:
                pass
        raise


def _finalize_reserved_job(
    job_id,
    user_login: str,
    parsed: ParsedSpreadsheet,
    proposal: MappingProposal,
) -> dict:
    with database.cursor(write=True, actor=user_login) as cursor:
        cursor.execute(
            """
            UPDATE logo.agent_spreadsheet_job
               SET format_name = %s,
                   status = 'mapping_pending',
                   mapping_hash = %s,
                   mapping = %s
             WHERE id = %s AND user_login = %s
               AND status = 'mapping_processing'
               AND expires_at > now()
            RETURNING *
            """,
            (
                parsed.format_name,
                mapping_hash(proposal),
                Json({**proposal.model_dump(mode="json"), "_command_counts": _sheet_command_counts(parsed, proposal)}),
                job_id,
                user_login,
            ),
        )
        row = cursor.fetchone()
        if row is None:
            raise Conflict("Spreadsheet reservation is no longer active")
        return dict(row)


def _discard_reserved_job(job: Mapping[str, Any], user_login: str, settings) -> None:
    """Compensate a failed pre-mapping reservation without touching jobs ready for review."""

    with database.cursor(write=True, actor=user_login) as cursor:
        cursor.execute(
            """
            DELETE FROM logo.agent_spreadsheet_job
             WHERE id = %s AND user_login = %s
               AND status = 'mapping_processing'
            """,
            (job["id"], user_login),
        )
        removed = cursor.rowcount == 1
    if removed:
        _remove_private_file(_job_path(settings, job["storage_key"]))


def _owned_job(cursor, job_id, user_login: str, *, lock: bool = False):
    try:
        parsed_id = uuid.UUID(str(job_id))
    except (ValueError, TypeError):
        raise NotFound("Spreadsheet job not found") from None
    suffix = " FOR UPDATE" if lock else ""
    cursor.execute(
        f"""
        SELECT * FROM logo.agent_spreadsheet_job
         WHERE id = %s AND user_login = %s{suffix}
        """,
        (parsed_id, user_login),
    )
    row = cursor.fetchone()
    if row is None:
        raise NotFound("Spreadsheet job not found")
    return row


async def create_spreadsheet_job(
    session_id,
    user_login: str,
    data: bytes,
    filename: str,
    media_type: str,
    instruction: str,
    settings,
    *,
    mapping_client=None,
    mapping_semaphore=None,
) -> dict:
    # Persistent admission is first: an over-cap request cannot consume parser
    # time or provider quota, and concurrent processes see the reservation.
    reserved_job = await _reserve_job_cancellation_safe(
        session_id=session_id,
        user_login=user_login,
        data=data,
        filename=filename,
        media_type=media_type,
        settings=settings,
    )
    known_reservation = None
    try:
        limits = SpreadsheetLimits.from_settings(settings)
        parsed = await _joinable_to_thread(
            parse_spreadsheet,
            data,
            filename,
            limits,
        )
        proposal = known_mapping(parsed)
        if proposal is not None:
            known_reservation = await quotas.reserve_async(
                user_login=user_login,
                reserved_tokens=1,
                settings=settings,
            )
        if proposal is None:
            proposal = await propose_mapping(
                list(parsed.headers),
                list(parsed.rows),
                instruction,
                settings,
                client=mapping_client,
                user_login=user_login,
                semaphore=mapping_semaphore,
            )
        validate_mapping_headers(proposal, parsed.headers)
        if known_reservation is not None:
            await asyncio.shield(_joinable_to_thread(
                quotas.reconcile,
                known_reservation,
                input_tokens=0,
                output_tokens=0,
            ))
            known_reservation = None
        return await _joinable_to_thread(
            _finalize_reserved_job,
            reserved_job["id"],
            user_login,
            parsed,
            proposal,
        )
    except BaseException:
        try:
            if known_reservation is not None:
                await asyncio.shield(_joinable_to_thread(
                    quotas.reconcile,
                    known_reservation,
                    input_tokens=0,
                    output_tokens=0,
                ))
        finally:
            await asyncio.shield(_joinable_to_thread(
                _discard_reserved_job,
                reserved_job,
                user_login,
                settings,
            ))
        raise


def get_spreadsheet_job(job_id, user_login: str) -> dict:
    with database.cursor() as cursor:
        return dict(_owned_job(cursor, job_id, user_login))


def _validate_mapping_confirmation(
    job: Mapping[str, Any],
    mapping_revision: int,
    confirmed_mapping_hash: str,
) -> None:
    if (
        job["status"] not in {"mapping_pending", "mapping_confirmed"}
        or job["expires_at"] <= datetime.now(timezone.utc)
    ):
        raise Conflict("Spreadsheet mapping is no longer pending")
    if (
        int(job["mapping_revision"]) != mapping_revision
        or job["mapping_hash"] != confirmed_mapping_hash
    ):
        raise Conflict("Mapping confirmation is stale")


def _terminal_confirmation_result(
    job_id,
    user_login: str,
    mapping_revision: int,
    confirmed_mapping_hash: str,
) -> Optional[dict]:
    """Return the canonical result of an already-finished confirmation."""

    with database.cursor() as cursor:
        job = dict(_owned_job(cursor, job_id, user_login))
    if (
        int(job["mapping_revision"]) != mapping_revision
        or job["mapping_hash"] != confirmed_mapping_hash
        or job["status"] not in {"staged", "rejected"}
    ):
        return None
    if job["status"] == "staged":
        if job.get("change_set_id") is None:
            raise Conflict("Completed spreadsheet job has no linked change-set")
        _attach_spreadsheet_change_sets(job,user_login)
    job["resolutions"] = (job.get("mapping") or {}).get("_resolutions", [])
    return job



def _ensure_chunk_links(job_id, user_login, first_id, count, chunk_size):
    """Persist links before staging so interruption/retry cannot orphan a chunk."""
    with database.cursor(write=True,actor=user_login) as cursor:
        job = dict(_owned_job(cursor,job_id,user_login,lock=True))
        mapping = job["mapping"]
        existing = mapping.get("_change_set_ids")
        if existing:
            if len(existing)!=count or mapping.get("_chunk_size")!=chunk_size:
                raise Conflict("Spreadsheet chunk size changed during retry")
            return existing
        if job["status"] not in {"mapping_pending","mapping_confirmed"}:
            raise Conflict("Spreadsheet mapping is no longer pending")
        ids = [str(first_id)]
        for _ in range(count-1):
            ids.append(str(staging.insert_change_set(cursor,job["session_id"],user_login,24,origin="spreadsheet")["id"]))
        cursor.execute("UPDATE logo.agent_spreadsheet_job SET mapping=mapping || %s::jsonb WHERE id=%s AND user_login=%s",
                       (Json({"_change_set_ids":ids,"_chunk_size":chunk_size}),job_id,user_login))
        return ids


def _attach_spreadsheet_change_sets(job, user_login):
    ids = (job.get("mapping") or {}).get("_change_set_ids") or [str(job["change_set_id"])]
    job["change_set_ids"] = ids
    # Full cards are fetched one at a time by the UI; a large sheet must not return thousands of snapshots.
    job["change_set"] = staging.get_change_set(ids[0],user_login)
    return job

def _ensure_linked_change_set(
    job_id,
    user_login: str,
    mapping_revision: int,
    confirmed_mapping_hash: str,
) -> tuple[dict, Optional[dict]]:
    """Atomically persist one recoverable job-to-change-set link."""

    with database.cursor(write=True, actor=user_login) as cursor:
        job = dict(_owned_job(cursor, job_id, user_login, lock=True))
        _validate_mapping_confirmation(
            job,
            mapping_revision,
            confirmed_mapping_hash,
        )
        if job.get("change_set_id") is not None:
            cursor.execute(
                """
                SELECT * FROM logo.agent_change_set
                 WHERE id = %s AND user_login = %s
                """,
                (job["change_set_id"], user_login),
            )
            existing = cursor.fetchone()
            if (
                existing is None
                or existing["status"] != "pending"
                or existing["expires_at"] <= datetime.now(timezone.utc)
            ):
                detail = "Linked change-set is no longer pending"
                cursor.execute(
                    """
                    UPDATE logo.agent_spreadsheet_job
                       SET status = 'rejected', rejected_rows = %s
                     WHERE id = %s AND user_login = %s
                    RETURNING *
                    """,
                    (
                        Json([{"row": None, "detail": detail}]),
                        job["id"],
                        user_login,
                    ),
                )
                return dict(cursor.fetchone()), None
            return job, dict(existing)

        cursor.execute(
            """
            SELECT id FROM logo.agent_chat_session
             WHERE id = %s AND user_login = %s AND expires_at > now()
             FOR UPDATE
            """,
            (job["session_id"], user_login),
        )
        if cursor.fetchone() is None:
            raise NotFound("Chat session not found")
        change_set = staging.insert_change_set(
            cursor,
            job["session_id"],
            user_login,
            24,
            origin="spreadsheet",
        )
        cursor.execute(
            """
            UPDATE logo.agent_spreadsheet_job
               SET status = 'mapping_confirmed', change_set_id = %s
             WHERE id = %s AND user_login = %s
            RETURNING *
            """,
            (change_set["id"], job["id"], user_login),
        )
        return dict(cursor.fetchone()), change_set


def _mark_job_rejected(job_id, user_login: str, rejected: list[dict]) -> None:
    with database.cursor(write=True, actor=user_login) as cursor:
        cursor.execute(
            """
            UPDATE logo.agent_spreadsheet_job
               SET status = 'rejected', rejected_rows = %s
             WHERE id = %s AND user_login = %s
               AND status IN ('mapping_pending', 'mapping_confirmed')
            RETURNING change_set_id
            """,
            (Json(rejected), job_id, user_login),
        )
        row = cursor.fetchone()
        if row is not None and row["change_set_id"] is not None:
            cursor.execute(
                """
                UPDATE logo.agent_change_set
                   SET status = 'discarded', updated_at = now()
                 WHERE id = %s AND user_login = %s AND status = 'pending'
                """,
                (row["change_set_id"], user_login),
            )


def _read_private_upload(path: Path, maximum: int) -> bytes:
    descriptor = None
    try:
        descriptor = os.open(
            path,
            os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0),
        )
        metadata = os.fstat(descriptor)
        if not stat.S_ISREG(metadata.st_mode):
            raise Conflict("Stored spreadsheet is unavailable")
        with os.fdopen(descriptor, "rb") as private_file:
            descriptor = None
            data = private_file.read(maximum + 1)
    except (FileNotFoundError, NotADirectoryError, OSError):
        raise Conflict("Stored spreadsheet is unavailable") from None
    finally:
        if descriptor is not None:
            os.close(descriptor)
    if len(data) > maximum:
        raise Conflict("Stored spreadsheet exceeds its upload limit")
    return data


def confirm_spreadsheet_mapping(
    job_id,
    user_login: str,
    mapping_revision: int,
    confirmed_mapping_hash: str,
    max_items: int,
    settings,
) -> dict:
    with database.cursor() as cursor:
        job = dict(_owned_job(cursor, job_id, user_login))
    terminal = _terminal_confirmation_result(
        job_id,
        user_login,
        mapping_revision,
        confirmed_mapping_hash,
    )
    if terminal is not None:
        return terminal
    _validate_mapping_confirmation(
        job,
        mapping_revision,
        confirmed_mapping_hash,
    )
    path = _job_path(settings, job["storage_key"])
    try:
        data = _read_private_upload(
            path,
            SpreadsheetLimits.from_settings(settings).max_bytes,
        )
        if hashlib.sha256(data).hexdigest() != job["sha256"]:
            raise Conflict(
                "Stored spreadsheet no longer matches the uploaded file"
            )
        parsed = parse_spreadsheet(
            data,
            job["original_name"],
            SpreadsheetLimits.from_settings(settings),
        )
        proposal = MappingProposal.model_validate({k:v for k,v in job["mapping"].items() if not k.startswith("_")})
        resolutions = []
        with database.cursor() as cursor:
            numbered_commands, rejected = _translate_rows_with_numbers(
                parsed, proposal, resolver=SpreadsheetNameResolver(cursor), resolutions=resolutions,
            )
    except (Conflict, InvalidCommand, ValueError) as exc:
        # A concurrent confirmer can finish and remove the private upload
        # between our initial read and file open. Return its canonical result
        # instead of turning an already-successful confirmation into a 409.
        terminal = _terminal_confirmation_result(
            job_id,
            user_login,
            mapping_revision,
            confirmed_mapping_hash,
        )
        if terminal is not None:
            return terminal
        _mark_job_rejected(
            job["id"],
            user_login,
            [{"row": None, "detail": str(exc)[:500]}],
        )
        _remove_private_file(path)
        if isinstance(exc, Conflict):
            raise
        raise InvalidCommand(str(exc)) from None

    max_items = max(1,min(int(max_items),int(getattr(settings,"agent_max_change_set_items",max_items)),staging.MAX_PERSISTED_CHANGE_SET_ITEMS))

    linked_job, change_set = _ensure_linked_change_set(
        job_id,
        user_login,
        mapping_revision,
        confirmed_mapping_hash,
    )
    if change_set is None:
        _remove_private_file(path)
        raise Conflict("Linked change-set is no longer pending")
    chunks = [numbered_commands[i:i+max_items] for i in range(0,len(numbered_commands),max_items)] or [[]]
    change_set_ids = _ensure_chunk_links(linked_job["id"],user_login,change_set["id"],len(chunks),max_items)
    staged_count = 0
    for chunk_id, chunk in zip(change_set_ids,chunks):
        source_rows_by_call_id = {f"spreadsheet:{linked_job['id']}:{row}": row for row,_command in chunk}
        batch_result = staging.stage_write_batch(
            chunk_id,
            [(next(name for name,model in COMMAND_MODELS.items() if type(command) is model),
              command.model_dump(mode="json"),f"spreadsheet:{linked_job['id']}:{row}") for row,command in chunk],
            user_login,max_items=max_items,
        )
        for rejected_item in batch_result.get("rejected_items",[]):
            rejected.append({"row":source_rows_by_call_id.get(rejected_item.get("call_id")),
                             "detail":str(rejected_item.get("detail","Invalid row"))[:500]})
        staged_count += int(batch_result.get("items",0))

    final_status = "staged" if staged_count else "rejected"
    with database.cursor(write=True, actor=user_login) as job_cursor:
        locked = dict(_owned_job(
            job_cursor,
            job_id,
            user_login,
            lock=True,
        ))
        if locked.get("change_set_id") != change_set["id"]:
            raise Conflict("Spreadsheet job link changed concurrently")
        if locked["status"] in {"staged", "rejected"}:
            final_job = locked
        else:
            if locked["status"] not in {"mapping_confirmed", "mapping_pending"}:
                raise Conflict("Spreadsheet mapping is no longer pending")
            job_cursor.execute(
                """
                UPDATE logo.agent_spreadsheet_job
                   SET status = %s, rejected_rows = %s,
                       mapping = mapping || %s::jsonb
                 WHERE id = %s AND user_login = %s
                RETURNING *
                """,
                (
                    final_status,
                    Json(rejected[: settings.agent_max_spreadsheet_rows]),
                    Json({"_resolutions": resolutions}),
                    locked["id"],
                    user_login,
                ),
            )
            final_job = dict(job_cursor.fetchone())

    _remove_private_file(path)
    if final_job is not None:
        final_job["resolutions"] = (final_job.get("mapping") or {}).get("_resolutions", [])
    if final_job is None or change_set is None:
        raise RuntimeError("Spreadsheet staging did not produce a final result")
    if final_job["status"] == "rejected":
        try:
            staging.discard_change_set(change_set["id"], user_login)
        except Exception:
            pass
        return final_job
    return _attach_spreadsheet_change_sets(final_job,user_login)
